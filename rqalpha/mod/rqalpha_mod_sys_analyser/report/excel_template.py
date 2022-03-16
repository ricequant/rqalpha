import re
import os
import datetime
from copy import copy
from itertools import product
from typing import Any, Dict, Tuple, List, Type, Optional

import openpyxl
from openpyxl.styles.cell_style import StyleArray
import numpy
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


VALUE_NAME_RE = r"^#(?P<name>[a-z0-9_]+)#$"


class SheetSchema:
    def __init__(self, template_ws: Worksheet):
        self._cell_map: Dict[str, Tuple[int, int, StyleArray]] = {}
        for row, column in product(range(1, template_ws.max_row + 1), range(1, template_ws.max_column + 1)):
            cell = template_ws.cell(row, column)
            value = cell.value
            if value:
                result = re.match(VALUE_NAME_RE, value)
                if result:
                    name = result.groupdict()["name"]
                    self._cell_map[name] = (row, column, cell._style)

    def fill_worksheet(self, ws: Worksheet, data: Dict[str, Any]):
        raise NotImplementedError

    def _write_cell(self, ws: Worksheet, row: int, col: int, data: Any, style: Optional[StyleArray] = None):
        if data is None:
            data = numpy.nan
        if isinstance(data, datetime.date):
            if data is pd.NaT:
                data = ""
        cell = ws.cell(row, col, data)
        if style:
            cell._style = style


class ExcelTemplate:
    SCHEMA_CLASSES: Dict[str, Type[SheetSchema]]
    TEMPLATE_NAME: str

    def __init__(self):
        template_path = os.path.join(os.path.dirname(__file__), "templates", self.TEMPLATE_NAME + ".xlsx")
        self._template_wb = openpyxl.load_workbook(template_path)
        self._sheet_schemas: Dict[str, SheetSchema] = {
            ws.title: self.SCHEMA_CLASSES[ws.title](ws) for ws in self._template_wb.worksheets
        }

    def new_workbook(self, data: Dict[str, Dict[str, Any]], filename):
        wb = copy(self._template_wb)
        for key, schema in self._sheet_schemas.items():
            schema.fill_worksheet(wb[key], data.get(key, {}))
        wb.save(filename)


class SingleCellSchema(SheetSchema):
    def fill_worksheet(self, ws: Worksheet, data: Dict[str, Any]):
        for key, (row, column, _) in self._cell_map.items():
            self._write_cell(ws, row, column, data.get(key))


class VerticalSeriesSchema(SheetSchema):
    def fill_worksheet(self, ws: Worksheet, data: Dict[str, List]):
        if not data:
            for key, (row, column, style) in self._cell_map.items():
                self._write_cell(ws, row, column, None, style)
        for key, (row, column, style) in self._cell_map.items():
            for i, item in enumerate(data.get(key, [])):
                self._write_cell(ws, row + i, column, item, style)


class SummaryTemplate(ExcelTemplate):
    TEMPLATE_NAME = "summary"
    SCHEMA_CLASSES = {
        "概览": SingleCellSchema,
        "年度指标": VerticalSeriesSchema,
        "月度收益": VerticalSeriesSchema,
        "月度主动收益": VerticalSeriesSchema
    }


SUMMARY_TEMPLATE = SummaryTemplate()


def generate_xlsx_reports(data, output_path):
    SUMMARY_TEMPLATE.new_workbook(data, os.path.join(output_path, "summary.xlsx"))
